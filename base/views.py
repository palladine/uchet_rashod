from django.utils import formats
from django.views import View
from django.shortcuts import render, HttpResponse, HttpResponseRedirect
from django.urls import reverse
from .forms import (LoginForm, AddPostofficeForm, AddCartridgeForm, AddCartridgesFileForm, AddPartForm,
                    AddSupplyForm, ShowCartridgesForm, AddPartsFileForm, AddOPSForm, AddOPSForm_U, AddOPSFileForm,
                    AddSupplyOPSForm, AddPartOPSForm, AddUserForm, AddGroupForm, AddPostofficeGroupForm, ShowOPSForm,
                    ShowRefuseForm, ShowRefuseFormUser, AutoorderFormPartChange, AutoorderFormNewPart, ShowOrderForm,
                    ChangeAmountAutoorderForm)
from django.contrib.auth import authenticate, login, logout
from .models import (User, Postoffice, Cartridge, Supply, Part, State, OPS, Supply_OPS, Part_OPS, State_OPS, Act,
                     Act_Postoffice, Group, AutoOrder, Part_AutoOrder, Supply_Stock, Part_Stock)
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.db.models import F, Sum, Count
import pandas
import openpyxl as xl
from openpyxl.styles import Alignment, Side, Border, Font
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
import os
from datetime import datetime, timedelta
import mimetypes
from acc_materials.settings import BASE_DIR, STATIC_ROOT



def get_errors_form(form):
    errors_dict = {}
    for field in form.errors:
        # field -> string
        for f in form:
            if field == f.name:
                errors_dict[f.label] = form.errors[field].as_text()

    errors_list = ['{}'.format(v) for k, v in errors_dict.items()]
    errors_str = "; ".join(errors_list).replace('*', '').replace('.', '')
    return errors_str



class UserLogin(View):
    def get(self, request):
        if not request.user.is_authenticated:
            form = LoginForm()
            context = {'form': form}
            return render(request, 'login.html', context=context)
        else:
            return HttpResponseRedirect(reverse('main'))


    def post(self, request):
        form = LoginForm(request.POST)
        if form.is_valid():
            username = request.POST['login']
            password = request.POST['password']

            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return HttpResponseRedirect(reverse('main'))

            else:
                error ='Неверное имя пользователя или пароль'
                context = {'form': form, 'error': error}
                return render(request, 'login.html', context=context)

        else:
            messages.error(request, get_errors_form(form))

            context = {'form': form}
            return render(request, 'login.html', context=context)



class UserLogout(View):
    def get(self, request):
        logout(request)
        return HttpResponseRedirect(reverse('login'))



class Main(View):
    def get(self, request):
        if request.user.is_authenticated:
            user = request.user
            context = {'user': user,
                       'title': 'Главная'}

            postoffice = user.postoffice_id.postoffice_name
            if user.role == '2':
                request.session['num_active_supplies'] = Supply.objects.filter(
                    postoffice_recipient=postoffice,
                    status_sending=True,
                    status_receiving=False).count()
                context.update({'num_active_supplies': request.session['num_active_supplies']})

            if user.role == '1':
                request.session['num_active_orders'] = AutoOrder.objects.filter(status_sending=True, viewed=False).count()
                context.update({'num_active_orders': request.session['num_active_orders']})

            return render(request, 'main.html', context=context)
        else:
            return HttpResponseRedirect(reverse('login'))



class AddGroup(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {}
        user = request.user

        form = AddGroupForm()

        context.update({'user': user, 'title': 'Добавить группу', 'form': form, 'num_active_orders': request.session['num_active_orders']})
        return render(request, 'addgroup.html', context=context)


    def post(self, request):
        context = {}
        form = AddGroupForm(request.POST)
        context.update({'title': 'Добавление группы', 'form': form, 'num_active_orders': request.session['num_active_orders']})

        if form.is_valid():
            group_name = request.POST.get('group_name')

            # save
            group = Group(group_name=group_name)
            group.save()
            messages.success(request, 'Группа добавлена')
            return HttpResponseRedirect(reverse('add_group'))
        else:
            messages.error(request, get_errors_form(form))
        return render(request, 'addgroup.html', context=context)



class ShowUsers(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        user = request.user

        context = {'user': user, 'title': 'Зарегистрированные пользователи'}

        if user.role == '1':
            if user.is_staff:
                all_users = User.objects.all().order_by('username')
            else:
                all_users = User.objects.filter(group=user.group).order_by('username')

            context.update({'users': all_users, 'num_active_orders': request.session['num_active_orders']})

        return render(request, 'showusers.html', context=context)



class AddPostoffice(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {}
        user = request.user

        form = AddPostofficeForm()
        if user.is_staff:
            form = AddPostofficeGroupForm()

        context.update({'user': user, 'title': 'Добавление почтамта', 'form': form, 'num_active_orders': request.session['num_active_orders']})
        return render(request, 'addpostoffice.html', context=context)


    def post(self, request):
        context = {}
        user = request.user
        form = AddPostofficeForm(request.POST)
        if user.is_staff:
            form = AddPostofficeGroupForm(request.POST)
        context.update({'user': user, 'title': 'Добавление почтамта', 'form': form, 'num_active_orders': request.session['num_active_orders']})

        if form.is_valid():
            group = user.group
            if user.is_staff:
                group_name = request.POST.get('group')
                group = Group.objects.get(group_name=group_name)
            postoffice_name = request.POST.get('postoffice_name')
            index = request.POST.get('index')
            address = request.POST.get('address')

            # save
            postoffice = Postoffice(postoffice_name=postoffice_name, index=index, address=address, group=group)
            postoffice.save()
            messages.success(request, 'Почтамт добавлен')
            return HttpResponseRedirect(reverse('add_postoffice'))
        else:
            messages.error(request, get_errors_form(form))
        return render(request, 'addpostoffice.html', context=context)



class AddOPS(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {}
        user = request.user

        context.update({'user': user, 'title': 'Добавление ОПС'})

        if user.role == '1':
            form_ops = AddOPSForm(user)
            context.update({'form': form_ops, 'num_active_orders': request.session['num_active_orders']})

        if user.role == '2':
            form_ops = AddOPSForm_U()
            context.update({'form': form_ops, 'num_active_supplies': request.session['num_active_supplies']})

        form_file_ops = AddOPSFileForm()
        context.update({'form_multy_ops': form_file_ops})

        return render(request, 'addops.html', context=context)


    def post(self, request):
        context = {}
        user = request.user
        context.update({'title': 'Добавление ОПС'})

        form_ops = False

        if user.role == '1':
            form_ops = AddOPSForm(user, request.POST)
            context.update({'form': form_ops, 'num_active_orders': request.session['num_active_orders']})

        if user.role == '2':
            form_ops = AddOPSForm_U(request.POST)
            context.update({'form': form_ops})

        form_file_ops = AddOPSFileForm(request.POST, request.FILES)

        if 'single_ops' in request.POST:
            if form_ops.is_valid():

                postoffice_name = request.POST.get('postoffice_name').split(' [')[0] if user.role == "1" else user.postoffice_id.postoffice_name
                index = request.POST.get('index')
                address = request.POST.get('address')

                # save
                postoffice = Postoffice.objects.get(postoffice_name=postoffice_name)
                # ops = OPS(postoffice=postoffice, index=index, address=address)

                ops, created_ops = OPS.objects.get_or_create(index=index)
                if created_ops:
                    ops.postoffice = postoffice
                    ops.address = address
                    ops.save()
                    messages.success(request, 'ОПС добавлено', extra_tags='single_ops')
                    return HttpResponseRedirect(reverse('add_ops'))
                else:
                    messages.error(request, "ОПС с таким индексом уже зарегистрировано", extra_tags='single_ops')

            else:
                messages.error(request, get_errors_form(form_ops), extra_tags='single_ops')


        if 'multy_ops' in request.POST:
            if form_file_ops.is_valid():
                file_object = form_file_ops.cleaned_data.get('file')
                file_bytes = file_object.read()

                wb = pandas.read_excel(file_bytes)  # wb -----> DataFrame

                count_created = 0
                for unit in wb.to_numpy():
                    postoffice_name = unit[0].strip()
                    index_ops = str(unit[1]).strip()
                    address = unit[2].strip()

                    postoffice = None
                    # admin
                    if user.role == '1':
                        postoffice = Postoffice.objects.filter(postoffice_name=postoffice_name, group=user.group).first()


                    # user
                    if user.role == '2':
                        if user.postoffice_id == Postoffice.objects.filter(postoffice_name=postoffice_name).first():
                            postoffice = user.postoffice_id
                        else:
                            continue

                    if postoffice:
                        obj, created = OPS.objects.get_or_create(index=index_ops, defaults={'postoffice': postoffice, 'address': address})
                        if created:
                            count_created += 1

                messages.success(request, f'ОПС добавлены ({count_created} позиций)', extra_tags='multy_ops')
                return HttpResponseRedirect(reverse('add_ops'))

            else:
                messages.error(request, get_errors_form(form_file_ops), extra_tags='multy_ops')


        if 'download_template_ops' in request.POST:
                try:
                    path = os.path.join(STATIC_ROOT, 'misc/')
                    filename = 'template_ops.xlsx'
                    file = open(path+filename, 'rb')
                    mime_type, _ = mimetypes.guess_type(path+filename)
                    response = HttpResponse(file, content_type=mime_type)
                    response['Content-Disposition'] = 'attachment; filename={}'.format(filename)

                    messages.success(request, '', extra_tags='download_template_ops')
                    return response
                except Exception as e:
                    messages.error(request, 'Ошибка скачивания файла ({})'.format(e), extra_tags='download_template_ops')


        return render(request, 'addops.html', context=context)




class AddCartridge(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {}
        user = request.user
        context.update({'user': user, 'title': 'Добавление номенклатуры картриджа', 'num_active_orders': request.session['num_active_orders']})

        if user.role != '1':
            return render(request, 'main.html', context=context)

        form_single = AddCartridgeForm()
        form_multy = AddCartridgesFileForm()

        context.update({'form_single': form_single, 'form_multy': form_multy})
        return render(request, 'addcartridge.html', context=context)


    def post(self, request):
        context = {}

        form_single = AddCartridgeForm(request.POST)
        form_multy = AddCartridgesFileForm(request.POST, request.FILES)
        context.update({'title': 'Добавление номенклатуры картриджа',
                        'form_single': form_single, 'form_multy': form_multy,
                        'num_active_orders': request.session['num_active_orders']})

        if 'single' in request.POST:
            if form_single.is_valid():
                nomenclature = request.POST.get('nomenclature').strip()
                printer = request.POST.get('printer_model')
                source = request.POST.get('source')
                drum = True if request.POST.get('is_drum') else False

                # save
                cartridge = Cartridge(nomenclature=nomenclature, printer_model=printer, is_drum=drum, source=source)
                cartridge.save()
                messages.success(request, 'Номенклатура картриджа добавлена', extra_tags='single')

                return HttpResponseRedirect(reverse('add_cartridge'))

            else:
                messages.error(request, get_errors_form(form_single), extra_tags='single')

        if 'multy' in request.POST:
            if form_multy.is_valid():
                file_object = form_multy.cleaned_data.get('file')
                file_bytes = file_object.read()

                wb = pandas.read_excel(file_bytes)  #  wb -----> DataFrame

                count_created = 0
                for unit in wb.to_numpy():
                    nomenclature = unit[0].strip()
                    printer = unit[1]
                    drum = True if str(unit[2]).strip() == 'Да' or unit[2] == 1 else False
                    source = '' if str(unit[3]) == 'nan' else str(unit[3])


                    # save
                    obj, created = Cartridge.objects.filter(nomenclature=nomenclature).get_or_create(
                        nomenclature=nomenclature,
                        printer_model=printer,
                        is_drum=drum,
                        source=source)
                    if created:
                        count_created += 1


                messages.success(request, f'Номенклатуры картриджей добавлены ({count_created} позиций)', extra_tags='multy')
                return HttpResponseRedirect(reverse('add_cartridge'))
            else:
                messages.error(request, get_errors_form(form_multy), extra_tags='multy')


        if 'download_template_cartridges' in request.POST:
            try:
                path = os.path.join(STATIC_ROOT, 'misc/')
                filename = 'template_cartridges.xlsx'
                file = open(path+filename, 'rb')
                mime_type, _ = mimetypes.guess_type(path+filename)
                response = HttpResponse(file, content_type=mime_type)
                response['Content-Disposition'] = 'attachment; filename={}'.format(filename)

                messages.success(request, '', extra_tags='download_template_cartridges')
                return response
            except Exception as e:
                messages.error(request, 'Ошибка скачивания файла ({})'.format(e), extra_tags='download_template_cartridges')

        return render(request, 'addcartridge.html', context=context)




class AddSupply(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {}
        user = request.user
        context.update({'user': user, 'title': 'Создать поставку картриджей на почтамт',
                        'num_active_orders': request.session['num_active_orders']})

        if user.role != '1':
            return render(request, 'main.html', context=context)

        form_supply = AddSupplyForm(user)
        form_part = AddPartForm(user)
        form_file_parts = AddPartsFileForm()

        postoffices = Postoffice.objects.filter(group=user.group)
        postoffices_names = [name.postoffice_name for name in postoffices]

        supplies = Supply.objects.filter(status_sending=False, postoffice_recipient__in=postoffices_names)
        active_supplies = [k for k in supplies if Part.objects.filter(id_supply=k.pk).count() > 0]
        ids = [i.pk for i in active_supplies]
        all_sup_wparts = [(Supply.objects.get(pk=p), Part.objects.filter(id_supply=p)) for p in ids]

        context.update({'form_part': form_part,
                        'form_supply': form_supply,
                        'form_file_parts': form_file_parts,
                        'supplies': all_sup_wparts})
        return render(request, 'addsupply.html', context=context)


    def post(self, request):
        context = {}
        user = request.user

        form_supply = AddSupplyForm(user, request.POST)
        form_part = AddPartForm(user, request.POST)
        form_file_parts = AddPartsFileForm(request.POST, request.FILES)

        context.update(
            {'title': 'Создать поставку картриджей на почтамт',
             'form_supply': form_supply,
             'form_part': form_part,
             'form_file_parts': form_file_parts,
             'user': user,
             'num_active_orders': request.session['num_active_orders']})

        if 'but_supply' in request.POST:
            if form_supply.is_valid():
                postoffice = request.POST.get('postoffice_name')

                # save supply
                supply = Supply(postoffice_recipient=postoffice.split(' [')[0])
                supply.save()

                messages.success(request,
                                 '{} создана. Необходимо добавить позиции в поставку.'.format(supply),
                                 extra_tags='supply')
                return HttpResponseRedirect(reverse('add_supply'))
            else:
                messages.error(request, get_errors_form(form_supply), extra_tags='supply')


        if 'but_part' in request.POST:
            if form_part.is_valid():
                # id_supply = request.POST.get('supply')
                id_supply = request.POST.get('new_supply')
                supply = Supply.objects.get(pk=id_supply)

                nomenclature = request.POST.get('nomenclature_cartridge')
                active_cartridge = Cartridge.objects.get(nomenclature=nomenclature)
                amount = request.POST.get('amount')

                a_sklad = Postoffice.objects.get(group=user.group, as_base=True)
                state = State.objects.get(postoffice=a_sklad, cartridge=active_cartridge)

                if int(amount) <= state.total_amount:
                # save Part
                    part = Part(id_supply=id_supply, postoffice=supply.postoffice_recipient, cartridge=nomenclature, amount=amount)
                    part.save()

                    # change state
                    state.blocked_amount += int(part.amount)
                    state.total_amount -= int(part.amount)
                    state.save()


                    messages.success(request,
                                     'Позиция в поставку ({}) добавлена.'.format(supply), extra_tags='part')
                    return HttpResponseRedirect(reverse('add_supply'))
                else:
                    messages.error(request,
                                   f"Запрошенного количества картриджей ({nomenclature}) нет на складе. "
                                   f"Имеется {state.total_amount}.", extra_tags='part')

                    form_part = AddPartForm(user)
                    context.update({'form_part': form_part})


            else:
                messages.error(request, get_errors_form(form_part), extra_tags='part')


        if 'but_file_supply' in request.POST:
            if form_file_parts.is_valid():
                file_object = form_file_parts.cleaned_data.get('file')
                file_bytes = file_object.read()

                wb = pandas.read_excel(file_bytes)  # wb -----> DataFrame
                wb = wb.dropna()

                nomenclatures = [x.nomenclature for x in Cartridge.objects.all()]
                postoffices = [x.postoffice_name for x in Postoffice.objects.all()]

                all_errors = []
                all_active_parts = []
                for unit in wb.to_numpy():

                    nomenclature = str(unit[0]).strip()
                    postoffice = str(unit[1]).strip()
                    amount = str(unit[2]).strip()

                    amount = int(amount) if amount.isdigit() else False

                    msg_error = ''
                    if (nomenclature in nomenclatures) and (postoffice in postoffices) and (amount > 0):
                        all_active_parts.append([nomenclature, postoffice, amount])

                    if (nomenclature not in nomenclatures):
                        msg_error += 'Неправильная номенклатура картриджа "{}"; '.format(nomenclature)

                    if (postoffice not in postoffices):
                        msg_error += 'Неправильное имя почтамта "{}"; '.format(postoffice)

                    if (not amount):
                        msg_error += 'Поле количество должно быть числом; '

                    if msg_error:
                        all_errors.append(msg_error)


                # TODO: output all errors !
                # !!! errors
                # print(all_errors)

                rel_post_sup = dict()
                for act_part in all_active_parts:
                    po = act_part[1]
                    if po in rel_post_sup:
                        # create part
                        part = Part(id_supply=rel_post_sup[po], postoffice=po, cartridge=act_part[0], amount=act_part[2])
                        part.save()
                    else:
                        # create supply
                        active_supply = Supply(postoffice_recipient=po)
                        active_supply.save()

                        rel_post_sup[po] = active_supply.pk

                        # create part
                        part = Part(id_supply=active_supply.pk, postoffice=po, cartridge=act_part[0], amount=act_part[2])
                        part.save()


        if 'download_template_parts' in request.POST:
            try:
                path = os.path.join(STATIC_ROOT, 'misc/')
                filename = 'template_parts.xlsx'
                file = open(path + filename, 'rb')
                mime_type, _ = mimetypes.guess_type(path + filename)
                response = HttpResponse(file, content_type=mime_type)
                response['Content-Disposition'] = 'attachment; filename={}'.format(filename)

                messages.success(request, '', extra_tags='download_template_parts')
                return response
            except Exception as e:
                messages.error(request, 'Ошибка скачивания файла ({})'.format(e), extra_tags='download_template_parts')


        id_sup_send = False
        for var in request.POST:
            if var.startswith('butsend'):
                id_sup_send = var.split('_')[1]

        if id_sup_send:
            user = request.user
            user_sender = '{0} ({1} {2})'.format(user.username, user.last_name, user.first_name)

            parts_send = Part.objects.filter(id_supply=id_sup_send)
            data_text = ''
            for p in parts_send:
                str_act = '{0}:{1};'.format(p.cartridge, p.amount)
                data_text += str_act

            date_sending = datetime.now()
            status_sending = True

            Supply.objects.filter(pk=id_sup_send).update(user_sender=user_sender,
                                                        data_text=data_text,
                                                        date_sending=date_sending,
                                                        status_sending=status_sending)
            return HttpResponseRedirect(reverse('add_supply'))


        id_part_del = False
        for var in request.POST:
            if var.startswith('butpartdel'):
                id_part_del = var.split('_')[1]

        if id_part_del:
            part_del = Part.objects.get(pk=id_part_del)

            # change state
            a_sklad = Postoffice.objects.get(group=user.group, as_base=True)
            pd_cartridge = Cartridge.objects.get(nomenclature=part_del.cartridge)
            state = State.objects.get(postoffice=a_sklad, cartridge=pd_cartridge)
            state.blocked_amount -= part_del.amount
            state.total_amount += part_del.amount
            state.save()

            #Part.objects.filter(pk=id_part_del).delete()
            part_del.delete()
            return HttpResponseRedirect(reverse('add_supply'))

        id_sup_del = False
        for var in request.POST:
            if var.startswith('butsupdel'):
                id_sup_del = var.split('_')[1]

        if id_sup_del:
            sup_del = Supply.objects.get(pk=id_sup_del)
            parts_del = Part.objects.filter(id_supply=id_sup_del)
            a_sklad = Postoffice.objects.get(group=user.group, as_base=True)


            for part_del in parts_del:
                pd_cartridge = Cartridge.objects.get(nomenclature=part_del.cartridge)
                state = State.objects.get(postoffice=a_sklad, cartridge=pd_cartridge)
                state.blocked_amount -= part_del.amount
                state.total_amount += part_del.amount
                state.save()

            parts_del.delete()
            sup_del.delete()
            return HttpResponseRedirect(reverse('add_supply'))


        postoffices = Postoffice.objects.filter(group=user.group)
        postoffices_names = [name.postoffice_name for name in postoffices]

        supplies = Supply.objects.filter(status_sending=False, postoffice_recipient__in=postoffices_names)
        active_supplies = [k for k in supplies if Part.objects.filter(id_supply=k.pk).count() > 0]
        ids = [i.pk for i in active_supplies]
        all_sup_wparts = [(Supply.objects.get(pk=p), Part.objects.filter(id_supply=p)) for p in ids]

        context.update({'supplies': all_sup_wparts})

        return render(request, 'addsupply.html', context=context)


## Принятие поставки на почтамте
class ApplySupply(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {}
        user = request.user

        if user.role == '1':
            return render(request, 'main.html', context=context)

        postoffice = user.postoffice_id
        supplies = Supply.objects.filter(postoffice_recipient=postoffice, status_sending=True, status_receiving=False)
        ids = [i.pk for i in supplies]
        all_sup_wparts = [(Supply.objects.get(pk=p), Part.objects.filter(id_supply=p)) for p in ids]

        context.update({'supplies': all_sup_wparts})
        context.update({'num_active_supplies': request.session['num_active_supplies']})

        return render(request, 'applysupply.html', context=context)


    def post(self, request):
        context = {}

        id_but_apply = False
        for var in request.POST:
            if var.startswith('butapply'):
                id_but_apply = var.split('_')[1]

        if id_but_apply:
            supply = Supply.objects.get(pk=id_but_apply)
            user = request.user
            supply.user_recipient = '{0} ({1} {2})'.format(user.username, user.last_name, user.first_name)
            supply.date_receiving = datetime.now()
            supply.status_receiving = True


            # change States
            postoffice_apply = Postoffice.objects.get(postoffice_name=supply.postoffice_recipient)

            active_parts = Part.objects.filter(id_supply=supply.pk)
            for part in active_parts:
                    cartridge_apply = Cartridge.objects.get(nomenclature=part.cartridge)
                    state, state_created = State.objects.get_or_create(cartridge_id=cartridge_apply.pk, postoffice_id=postoffice_apply.pk)
                    if state_created:
                        state.total_amount = part.amount
                    else:
                        state.total_amount += part.amount
                    state.save()

                    active_group = user.postoffice_id.group
                    sklad = Postoffice.objects.get(group=active_group, as_base=True)
                    state_sklad = State.objects.get(postoffice=sklad, cartridge=cartridge_apply)
                    state_sklad.blocked_amount -= part.amount

                    state_sklad.save()

            supply.save()  # !!!! after all actions

            #####
            postoffice = user.postoffice_id.postoffice_name
            request.session['num_active_supplies'] = Supply.objects.filter(
                postoffice_recipient=postoffice,
                status_sending=True,
                status_receiving=False).count()


            return HttpResponseRedirect(reverse('apply_supply'))

        postoffice = request.user.postoffice_id
        supplies = Supply.objects.filter(postoffice_recipient=postoffice, status_sending=True, status_receiving=False)
        ids = [i.pk for i in supplies]
        all_sup_wparts = [(Supply.objects.get(pk=p), Part.objects.filter(id_supply=p)) for p in ids]
        context.update({'supplies': all_sup_wparts})
        context.update({'num_active_supplies': request.session['num_active_supplies']})

        return render(request, 'applysupply.html', context=context)



class ShowCartridges(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {'title': 'Картриджи на почтамте'}

        user = request.user

        states = request.session.get('states', False)
        postoffice = request.session.get('postoffice', False)

        if postoffice:
            request.session['postoffice'] = False

        if isinstance(states, list):
            request.session['states'] = False
        else:
            ### по аналогии с post !
            postoffice = user.postoffice_id
            postoffice_id = Postoffice.objects.get(postoffice_name=postoffice.postoffice_name.split(" [")[0]).pk
            query = State.objects.filter(postoffice_id=postoffice_id, total_amount__gte=0)
            states = []
            for q in query:
                states.append([q.cartridge.nomenclature, q.cartridge.printer_model, q.total_amount])
            #states.sort()
            states = sorted(states, key=lambda x: x[2], reverse=True)



        if user.role == '1':
            form = ShowCartridgesForm(user)
            context.update({'form': form, 'num_active_orders': request.session['num_active_orders']})

        if user.role == '2':
            context.update({'num_active_supplies': request.session['num_active_supplies']})

        context.update({'states': states, 'postoffice': postoffice})

        return render(request, 'showcartridges.html', context=context)


    def post(self, request):
        context = {}
        user = request.user

        form = ShowCartridgesForm(user, request.POST)
        context.update({'title': 'Картриджи на почтамте', 'form': form, 'num_active_orders': request.session['num_active_orders']})

        if form.is_valid():
            postoffice_name = request.POST.get('postoffice_name').split(" [")[0]
            postoffice_id = Postoffice.objects.get(postoffice_name=postoffice_name).pk

            query = State.objects.filter(postoffice_id=postoffice_id, total_amount__gte=0)

            states = []
            for q in query:
                states.append([q.cartridge.nomenclature, q.cartridge.printer_model, q.total_amount])
            #states.sort()
            states = sorted(states, key=lambda x: x[2], reverse=True)

            request.session['states'] = states
            request.session['postoffice'] = postoffice_name

            return HttpResponseRedirect(reverse('show_cartridges'))

        else:
            messages.error(request, get_errors_form(form))


        postoffice = request.user.postoffice_id.postoffice_name
        postoffice_id = Postoffice.objects.get(postoffice_name=postoffice).pk

        query = State.objects.filter(postoffice_id=postoffice_id, total_amount__gte=0)
        states = []
        for q in query:
            states.append([q.cartridge.nomenclature, q.cartridge.printer_model, q.total_amount])

        context.update({'states': states, 'postoffice': postoffice, 'form': form})
        return render(request, 'showcartridges.html', context=context)


class ShowNomenclatures(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {'title': 'Зарегистрированные номенклатуры картриджей'}

        user = request.user
        if user.role == '1':
            context.update({'num_active_orders': request.session['num_active_orders']})

        if user.role == '2':
            context.update({'num_active_supplies': request.session['num_active_supplies']})

        cartridges = Cartridge.objects.all().order_by('nomenclature')

        # pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(cartridges, 25)
        try:
            nomenclatures = paginator.page(page)
        except PageNotAnInteger:
            nomenclatures = paginator.page(1)
        except EmptyPage:
            nomenclatures = paginator.page(paginator.num_pages)

        context.update({'nomenclatures': nomenclatures})

        return render(request, 'shownomenclatures.html', context=context)


class ShowOPS(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {}
        user = request.user
        form = ShowOPSForm()

        opss = False

        if user.role == '1':
            postoffices = Postoffice.objects.filter(group=user.group)
            opss = OPS.objects.filter(postoffice__in=postoffices).order_by('postoffice__postoffice_name', 'index')

            if user.is_staff:
                ids = request.session.get('ops_ids', False)
                if ids:
                    opss = OPS.objects.filter(pk__in=ids).order_by('postoffice__postoffice_name', 'index')


                context.update({'form': form})

            context.update({'title': 'Список зарегистрированных ОПС', 'num_active_orders': request.session['num_active_orders']})


        if user.role == '2':
            context.update({'title': 'Список ОПС Почтамта', 'num_active_supplies': request.session['num_active_supplies']})
            postoffice_obj = Postoffice.objects.get(postoffice_name=user.postoffice_id.postoffice_name)
            opss = OPS.objects.filter(postoffice=postoffice_obj).order_by('index')

        # pagination
        if opss:
            page = request.GET.get('page', 1)
            paginator = Paginator(opss, 25)
            try:
                ops_all = paginator.page(page)
            except PageNotAnInteger:
                ops_all = paginator.page(1)
            except EmptyPage:
                ops_all = paginator.page(paginator.num_pages)

            request.session['ops_ids'] = False
            context.update({'ops_all': ops_all})

        return render(request, 'showops.html', context=context)


    def post(self, request):
        context = {}
        user = request.user

        form = ShowOPSForm(request.POST)
        context.update({'title': 'Список зарегистрированных ОПС', 'form': form, 'num_active_orders': request.session['num_active_orders']})

        if form.is_valid():
            group_name = request.POST.get('group')
            group = Group.objects.get(group_name=group_name)

            postoffices = Postoffice.objects.filter(group=group)
            opss = OPS.objects.filter(postoffice__in=postoffices).order_by('index')

            opsids = [o.pk for o in opss]
            request.session['ops_ids'] = opsids
            return HttpResponseRedirect(reverse('show_ops'))

        else:
            messages.error(request, get_errors_form(form))

        postoffices = Postoffice.objects.filter(group=user.group)
        opss = OPS.objects.filter(postoffice__in=postoffices).order_by('index')
        context.update({'opss': opss})
        return render(request, 'showops.html', context=context)


class AddSupplyOPS(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {}
        user = request.user

        if user.role == '1':
            return render(request, 'main.html', context=context)

        postoffice = user.postoffice_id

        form_supply_ops = AddSupplyOPSForm(postoffice)
        form_part_ops = AddPartOPSForm(postoffice)

        qops = OPS.objects.filter(postoffice=postoffice)
        supplies_ops = Supply_OPS.objects.filter(status_sending=False, ops_recipient__in=qops)
        active_supplies = [k for k in supplies_ops if Part_OPS.objects.filter(id_supply_ops=k.pk).count() > 0]
        ids = [i.pk for i in active_supplies]
        all_sup_wparts = [(Supply_OPS.objects.get(pk=p), Part_OPS.objects.filter(id_supply_ops=p)) for p in ids]

        context.update({'user': user,
                        'title': 'Создать поставку картриджей на опс',
                        'form_part_ops': form_part_ops,
                        'form_supply_ops': form_supply_ops,
                        'supplies_ops': all_sup_wparts,
                        'num_active_supplies': request.session['num_active_supplies']})

        return render(request, 'addsupplyops.html', context=context)


    def post(self, request):
        context = {}

        user = request.user
        postoffice = user.postoffice_id
        postoffice_obj = Postoffice.objects.get(postoffice_name=postoffice)

        form_supply_ops = AddSupplyOPSForm(postoffice, request.POST)
        form_part_ops = AddPartOPSForm(postoffice, request.POST)
        #form_file_parts = AddPartsFileForm(request.POST, request.FILES)


        context.update(
            {'title': 'Создать поставку картриджей на опс',
             'form_supply_ops': form_supply_ops,
             'form_part_ops': form_part_ops,
             'num_active_supplies': request.session['num_active_supplies'],
             #'form_file_parts': form_file_parts,
             'user': user})

        if 'but_supply' in request.POST:
            if form_supply_ops.is_valid():
                #ops = request.POST.get('ops')
                ops = request.POST.get('new_ops')  # js
                task_naumen = request.POST.get('task_naumen')
                ops_obj = OPS.objects.get(pk=ops)
                # save supply_ops
                supply_ops = Supply_OPS(ops_recipient=ops_obj, id_task_naumen=task_naumen)
                supply_ops.save()

                messages.success(request,
                                 '{} создана. Необходимо добавить позиции в поставку.'.format(supply_ops),
                                 extra_tags='supply')
                return HttpResponseRedirect(reverse('add_supply_ops'))
            else:
                messages.error(request, get_errors_form(form_supply_ops), extra_tags='supply')
                form_supply_ops = AddSupplyOPSForm(postoffice)
                context.update({'form_supply_ops': form_supply_ops})

        if 'but_part' in request.POST:
            if form_part_ops.is_valid():
                # id_supply = request.POST.get('supply_ops')
                id_supply = request.POST.get('new_supply_ops')  # js
                supply_obj = Supply_OPS.objects.get(pk=id_supply)

                nomenclature = request.POST.get('nomenclature_cartridge')
                cartridge_obj = Cartridge.objects.get(nomenclature=nomenclature)
                amount = request.POST.get('amount')

                #postoffice_obj = supply_obj.ops_recipient.postoffice

                state_amount = State.objects.filter(postoffice=postoffice_obj, cartridge=cartridge_obj).first()

                ####
                ops_postoffice = OPS.objects.filter(postoffice=postoffice_obj)
                open_supplies_postoffice = Supply_OPS.objects.filter(status_sending=False, ops_recipient__in=ops_postoffice)
                open_parts = Part_OPS.objects.filter(id_supply_ops__in=open_supplies_postoffice, cartridge=cartridge_obj)

                ## количество картриджей номенклатуры в открытых поставках
                parts_amount = open_parts.aggregate(Sum('amount')).get('amount__sum', 0)
                parts_amount = parts_amount if parts_amount else 0

                if int(amount) <= state_amount.total_amount-parts_amount:
                    # save Part
                    part_ops = Part_OPS(id_supply_ops=supply_obj, cartridge=cartridge_obj, amount=amount)
                    part_ops.save()

                    messages.success(request, 'Позиция в поставку ({}) добавлена.'.format(supply_obj.pk), extra_tags='part')
                    return HttpResponseRedirect(reverse('add_supply_ops'))
                else:
                    messages.error(request,
                                   f"Запрошенного количества картриджей ({nomenclature}) нет на почтамте (с учетом открытых поставок на ОПС). "
                                   f"Имеется {state_amount.total_amount - parts_amount}.", extra_tags='part')
                    form_part_ops = AddPartOPSForm(postoffice)
                    context.update({'form_part_ops': form_part_ops})
            else:
                messages.error(request, get_errors_form(form_part_ops), extra_tags='part')
                form_part_ops = AddPartOPSForm(postoffice)
                context.update({'form_part_ops': form_part_ops})


        id_sup_send = False
        for var in request.POST:
            if var.startswith('butsend'):
                id_sup_send = var.split('_')[1]

        if id_sup_send:
            user = request.user
            parts_send = Part_OPS.objects.filter(id_supply_ops__pk=id_sup_send)

            data_text = ''
            for part in parts_send:
                str_act = '{0}:{1};'.format(part.cartridge, part.amount)
                data_text += str_act

            date_sending = datetime.now()
            status_sending = True

            Supply_OPS.objects.filter(pk=id_sup_send).update(user_sender=user,
                                                         data_text=data_text,
                                                         date_sending=date_sending,
                                                         status_sending=status_sending)

            # change state ops
            for part in parts_send:
                state, state_created = State_OPS.objects.get_or_create(ops=part.id_supply_ops.ops_recipient, cartridge=part.cartridge)
                if state_created:
                    state.total_amount = part.amount
                else:
                    state.total_amount += part.amount
                state.save()

            # change state postoffice
            for part in parts_send:
                State.objects.filter(postoffice=postoffice_obj, cartridge=part.cartridge).update(total_amount=F('total_amount')-part.amount)

            return HttpResponseRedirect(reverse('add_supply_ops'))


        id_part_del = False
        for var in request.POST:
            if var.startswith('butpartdel'):
                id_part_del = var.split('_')[1]

        if id_part_del:
            Part_OPS.objects.filter(pk=id_part_del).delete()
            return HttpResponseRedirect(reverse('add_supply_ops'))

        id_sup_del = False
        for var in request.POST:
            if var.startswith('butsupdel'):
                id_sup_del = var.split('_')[1]

        if id_sup_del:
            supply_obj = Supply_OPS.objects.get(pk=id_sup_del)
            Part_OPS.objects.filter(id_supply_ops=supply_obj).delete()
            supply_obj.delete()
            return HttpResponseRedirect(reverse('add_supply_ops'))

        qops = OPS.objects.filter(postoffice=postoffice)
        supplies_ops = Supply_OPS.objects.filter(status_sending=False, ops_recipient__in=qops)
        active_supplies = [k for k in supplies_ops if Part_OPS.objects.filter(id_supply_ops__pk=k.pk).count() > 0]
        ids = [i.pk for i in active_supplies]
        all_sup_wparts = [(Supply_OPS.objects.get(pk=p), Part_OPS.objects.filter(id_supply_ops__pk=p)) for p in ids]

        context.update({'supplies_ops': all_sup_wparts})

        return render(request, 'addsupplyops.html', context=context)



class ShowSupplyOPS(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        user = request.user
        context = {'user': user, 'title': 'Реестр поставок на ОПС'}

        if user.role == "1":
            return render(request, 'main.html', context=context)


        query_supplies = Supply_OPS.objects.filter(ops_recipient__postoffice=user.postoffice_id, status_sending=True).order_by('-id')
        headers = ['№ поставки', 'Индекс ОПС', 'Отправитель', 'Данные поставки', 'Запрос Naumen', 'Дата отправки', 'Акт распечатан', '']

        supplies = []
        for query_supply in query_supplies:
            dt = "<br>".join(query_supply.data_text.split(';'))

            supply = [query_supply.id,
                      query_supply.ops_recipient.index,
                      "{}<br>({} {})".format(query_supply.user_sender.username, query_supply.user_sender.first_name, query_supply.user_sender.last_name),
                      dt,
                      query_supply.id_task_naumen,
                      query_supply.date_sending]


            act = Act.objects.filter(id_supply_ops=query_supply)
            if act:
                supply.append(act.first().status_act)
            else:
                supply.append('')

            supplies.append(supply)


        # pagination
        if supplies:
            page = request.GET.get('page', 1)
            paginator = Paginator(supplies, 10)
            try:
                supplies_all = paginator.page(page)
            except PageNotAnInteger:
                supplies_all = paginator.page(1)
            except EmptyPage:
                supplies_all = paginator.page(paginator.num_pages)


            context.update({'supplies': supplies_all,
                            'headers': headers,
                            'num_active_supplies': request.session['num_active_supplies']})

        return render(request, 'showsupplyops.html', context=context)


    def post(self, request):
        user = request.user
        context = {'user': user, 'title': 'Реестр поставок на ОПС'}

        id_sup_act = False
        for var in request.POST:
            if var.startswith('butact'):
                id_sup_act = var.split('_')[1]

        if id_sup_act:
            supply_obj = Supply_OPS.objects.get(id=id_sup_act)
            act_obj, act_created = Act.objects.get_or_create(id_supply_ops=supply_obj)
            if act_created:
                act_obj.date_creating = datetime.now()
                act_obj.status_act = False
                act_obj.save()

            path = os.path.join(STATIC_ROOT, 'misc/')
            path_acts = os.path.join(STATIC_ROOT, 'misc/acts/')
            filename = 'template_act.xlsx'
            new_filename = f"act_{act_obj.pk}_{act_obj.date_creating.date()}.xlsx"

            new_wb = None
            if os.path.exists(path_acts+new_filename):
                new_wb = xl.load_workbook(path_acts+new_filename)
            else:
                # copy to new excel file
                wb = xl.load_workbook(path+filename)
                new_wb = wb

                ws = new_wb.active
                ws['F2'] = act_obj.pk
                ws['H2'] = act_obj.date_creating.date().strftime("%d.%m.%Y")
                user_sender = supply_obj.user_sender
                ws['E4'] = "{} {} {}".format(user_sender.last_name, user_sender.first_name, user_sender.middle_name)
                ws['E5'] = f"ОПС {supply_obj.ops_recipient.index}"
                ws['H6'] = datetime.now().strftime("%d.%m.%Y %H:%M")
                ws['H8'] = supply_obj.id_task_naumen

                c = 0
                for pos in supply_obj.data_text.split(";"):
                    if pos:
                        point = 12+c
                        elist = pos.split(":")
                        ws.insert_rows(point)
                        cell_range = 'C{0}:G{0}'.format(point)
                        ws.merge_cells(cell_range)

                        brd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        fnt = Font(name="Times New Roman", size=11)
                        ws['B{}'.format(point)] = c+1
                        ws['B{}'.format(point)].alignment = Alignment(horizontal='center')
                        ws['B{}'.format(point)].border = brd
                        ws['B{}'.format(point)].font = fnt
                        ws['C{}'.format(point)] = elist[0]
                        ws['H{}'.format(point)] = elist[1]
                        ws['H{}'.format(point)].alignment = Alignment(horizontal='center')
                        ws['H{}'.format(point)].border = brd
                        ws['H{}'.format(point)].font = fnt

                        # border merge cell
                        start_cell, end_cell = cell_range.split(':')
                        start_coord = coordinate_from_string(start_cell)
                        start_row = start_coord[1]
                        start_col = column_index_from_string(start_coord[0])
                        end_coord = coordinate_from_string(end_cell)
                        end_row = end_coord[1]
                        end_col = column_index_from_string(end_coord[0])

                        for row in range(start_row, end_row + 1):
                            for col_idx in range(start_col, end_col + 1):
                                col = get_column_letter(col_idx)
                                ws['{}{}'.format(col, row)].border = brd
                                ws['{}{}'.format(col, row)].font = fnt
                        c += 1
                new_wb.save(filename=path_acts+new_filename)

            # open act file in system
            #os.system("start EXCEL.EXE {}".format(path_acts+new_filename))

            try:
                file = open(path_acts+new_filename, 'rb')
                mime_type, _ = mimetypes.guess_type(path_acts+new_filename)
                response = HttpResponse(file, content_type=mime_type)
                response['Content-Disposition'] = 'attachment; filename={}'.format(new_filename)
                # messages.success(request, '', extra_tags='download_template_cartridges')
                act_obj.status_act = True
                act_obj.save()
                return response
            except Exception as e:
                # messages.error(request, 'Ошибка скачивания файла ({})'.format(e), extra_tags='download_template_cartridges')
                ...


        if 'butsave' in request.POST:
            query_supplies = Supply_OPS.objects.filter(ops_recipient__postoffice=user.postoffice_id,
                                                       status_sending=True).order_by('id')
            headers = ['№ поставки', 'Индекс ОПС', 'Отправитель', 'Данные поставки', 'Запрос Naumen', 'Дата отправки',
                       'Акт распечатан']

            supplies = []
            for query_supply in query_supplies:
                dt = "<br>".join(query_supply.data_text.split(';'))

                supply = [query_supply.id,
                          query_supply.ops_recipient.index,
                          "{}<br>({} {})".format(query_supply.user_sender.username, query_supply.user_sender.first_name,
                                                 query_supply.user_sender.last_name),
                          dt,
                          query_supply.id_task_naumen,
                          query_supply.date_sending]


                act = Act.objects.filter(id_supply_ops=query_supply)
                if act:
                    supply.append(act.first().status_act)
                else:
                    supply.append('')

                supplies.append(supply)


            filename_supplies = 'supplies_ops.xlsx'

            new_wb = xl.Workbook()
            ws = new_wb.active

            for h in range(len(headers)):
                ws_cell = ws.cell(1, h+1)
                ws_cell.font = xl.styles.Font(bold=True)
                ws_cell.value = headers[h]
                column_dimensions = ws.column_dimensions[get_column_letter(h+1)]
                if h == 3:
                    column_dimensions.width = 30
                else:
                    column_dimensions.width = 20


            for sp in range(len(supplies)):
                for inx_sp in range(len(supplies[sp])):
                    ws_cell = ws.cell(sp+2, inx_sp+1)
                    ws_cell.value = str(supplies[sp][inx_sp]).replace('<br>', ' ')


            mime_type, _ = mimetypes.guess_type(filename_supplies)
            response = HttpResponse(content_type=mime_type)
            response['Content-Disposition'] = 'attachment; filename={}'.format(filename_supplies)

            new_wb.save(response)
            return response

        return HttpResponseRedirect(reverse('show_supply_ops'))



class AddUser(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {}
        user = request.user
        form = AddUserForm(user)
        context.update({'user': user, 'title': 'Добавление пользователя', 'form': form, 'num_active_orders': request.session['num_active_orders']})
        return render(request, 'adduser.html', context=context)


    def post(self, request):
        context = {}
        user = request.user
        form = AddUserForm(user, request.POST)
        context.update({'title': 'Добавление пользователя', 'form': form, 'num_active_orders': request.session['num_active_orders']})

        if form.is_valid():
            login = request.POST.get('login')
            password = request.POST.get('password')
            first_name = request.POST.get('firstname') # Имя
            last_name = request.POST.get('lastname')   # Фамилия
            middle_name = request.POST.get('middlename')
            postoffice_name = request.POST.get('postoffice').split(" [")[0]
            postoffice = Postoffice.objects.get(postoffice_name=postoffice_name)
            group = postoffice.group
            email = request.POST.get('email')

            # save
            new_user = User(username=login, email=email, last_name=last_name, first_name=first_name,
                            middle_name=middle_name, group=group, postoffice_id=postoffice, date_joined=datetime.now(),
                            last_login=datetime.now(), role='2', is_staff=False, is_active=True)
            new_user.set_password(password)  # !!! пароль защифрован - только так
            new_user.save()
            messages.success(request, 'Пользователь добавлен')

            return HttpResponseRedirect(reverse('add_user'))

        else:
            messages.error(request, get_errors_form(form))

        return render(request, 'adduser.html', context=context)


class ShowSupply(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        user = request.user
        context = {'user': user, 'title': 'Реестр поставок на почтамты', 'num_active_orders': request.session['num_active_orders']}

        if user.role != '1':
            return render(request, 'main.html', context=context)

        postoffices = Postoffice.objects.filter(group=user.group)
        postoffices_names = [name.postoffice_name for name in postoffices]

        query_supplies = Supply.objects.filter(status_sending=True, postoffice_recipient__in=postoffices_names).order_by('-id')

        if user.is_staff:
            query_supplies = Supply.objects.filter(status_sending=True).order_by('-id')

        supplies = []
        for query_supply in query_supplies:
            dt = "<br>".join(query_supply.data_text.split(';'))

            ds = query_supply.date_sending
            date_s = ds if ds else ''

            dr = query_supply.date_receiving
            date_r = dr if dr else ''

            act = Act_Postoffice.objects.filter(id_supply=query_supply.pk)
            status = ''
            if act.exists():
                status = act.first().status_act

            supply = [query_supply.id,
                      query_supply.postoffice_recipient,
                      query_supply.user_sender,
                      query_supply.user_recipient,
                      dt,
                      date_s,
                      date_r,
                      query_supply.status_receiving,
                      status]

            supplies.append(supply)

        # pagination
        if supplies:
            page = request.GET.get('page', 1)
            paginator = Paginator(supplies, 10)
            try:
                supplies_all = paginator.page(page)
            except PageNotAnInteger:
                supplies_all = paginator.page(1)
            except EmptyPage:
                supplies_all = paginator.page(paginator.num_pages)

            context.update({'supplies': supplies_all})

        return render(request, 'showsupply.html', context=context)

    def post(self, request):

        user = request.user
        context = {'user': user, 'title': 'Реестр поставок на почтамты', 'num_active_orders': request.session['num_active_orders']}


        supply_id = False
        for var in request.POST:
            if var.startswith('supply'):
                supply_id = var.split('_')[1]

        if supply_id:
            supply_obj = Supply.objects.get(id=supply_id)
            act_obj, act_created = Act_Postoffice.objects.get_or_create(id_supply=supply_obj)
            if act_created:
                act_obj.date_creating = datetime.now()
                act_obj.status_act = False
                act_obj.save()

            path = os.path.join(STATIC_ROOT, 'misc/')
            path_acts = os.path.join(STATIC_ROOT, 'misc/acts/')
            filename = 'template_act.xlsx'
            new_filename = f"act_{act_obj.pk}_{act_obj.date_creating.date()}.xlsx"

            new_wb = None
            if os.path.exists(path_acts + new_filename):
                new_wb = xl.load_workbook(path_acts + new_filename)
            else:
                # copy to new excel file
                wb = xl.load_workbook(path + filename)
                new_wb = wb

                ws = new_wb.active
                ws['F2'] = act_obj.pk
                ws['H2'] = act_obj.date_creating.date().strftime("%d.%m.%Y")
                username_sender = supply_obj.user_sender.split(' ')[0]
                user_sender = User.objects.get(username=username_sender)
                ws['E4'] = "{} {} {}".format(user_sender.last_name, user_sender.first_name, user_sender.middle_name)
                ws['E5'] = f"{supply_obj.postoffice_recipient}"
                ws['H6'] = datetime.now().strftime("%d.%m.%Y %H:%M")
                #ws['H8'] = supply_obj.id_task_naumen

                c = 0
                for pos in supply_obj.data_text.split(";"):
                    if pos:
                        point = 12 + c
                        elist = pos.split(":")
                        ws.insert_rows(point)
                        cell_range = 'C{0}:G{0}'.format(point)
                        ws.merge_cells(cell_range)

                        brd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                        fnt = Font(name="Times New Roman", size=11)
                        ws['B{}'.format(point)] = c + 1
                        ws['B{}'.format(point)].alignment = Alignment(horizontal='center')
                        ws['B{}'.format(point)].border = brd
                        ws['B{}'.format(point)].font = fnt
                        ws['C{}'.format(point)] = elist[0]
                        ws['H{}'.format(point)] = elist[1]
                        ws['H{}'.format(point)].alignment = Alignment(horizontal='center')
                        ws['H{}'.format(point)].border = brd
                        ws['H{}'.format(point)].font = fnt

                        # border merge cell
                        start_cell, end_cell = cell_range.split(':')
                        start_coord = coordinate_from_string(start_cell)
                        start_row = start_coord[1]
                        start_col = column_index_from_string(start_coord[0])
                        end_coord = coordinate_from_string(end_cell)
                        end_row = end_coord[1]
                        end_col = column_index_from_string(end_coord[0])

                        for row in range(start_row, end_row + 1):
                            for col_idx in range(start_col, end_col + 1):
                                col = get_column_letter(col_idx)
                                ws['{}{}'.format(col, row)].border = brd
                                ws['{}{}'.format(col, row)].font = fnt
                        c += 1
                new_wb.save(filename=path_acts + new_filename)


            try:
                file = open(path_acts+new_filename, 'rb')
                mime_type, _ = mimetypes.guess_type(path_acts+new_filename)
                response = HttpResponse(file, content_type=mime_type)
                response['Content-Disposition'] = 'attachment; filename={}'.format(new_filename)

                act_obj.status_act = True
                act_obj.save()
                return response
            except Exception as e:
                # messages.error(request, 'Ошибка скачивания файла ({})'.format(e), extra_tags='download_template_cartridges')
                ...

            #return HttpResponseRedirect(reverse('show_supply'))


        return render(request, 'showsupply.html', context=context)


class ShowRefuse(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        context = {}
        user = request.user

        parts_final = request.session.get('sr_parts', False)
        if parts_final:
            request.session['sr_parts'] = False
            context.update({'sr_parts': parts_final})

        if user.role == '1':
            form = ShowRefuseForm(user)
            context.update(({'num_active_orders': request.session['num_active_orders']}))
        else:
            form = ShowRefuseFormUser()
            context.update({'num_active_supplies': request.session['num_active_supplies']})

        context.update({'user': user, 'title': 'Реестр списанных картриджей', 'form': form})

        return render(request, 'showrefuse.html', context=context)



    def post(self, request):
        context = {}

        user = request.user
        params_for_supplies = {}
        params_for_opss = {}

        if user.role == '1':
            form = ShowRefuseForm(user, request.POST)
            context.update({'num_active_orders': request.session['num_active_orders']})

            if form.is_valid():
                postoffice_name = request.POST.get('postoffice', False)
                if postoffice_name:
                    postoffice_name = postoffice_name.split(" [")[0]
                    postoffices = Postoffice.objects.filter(postoffice_name=postoffice_name)
                else:
                    if user.is_staff:
                        postoffices = Postoffice.objects.all()
                    else:
                        postoffices = Postoffice.objects.filter(group=user.group)

                params_for_opss.update({'postoffice__in': postoffices})

            else:
                messages.error(request, get_errors_form(form))

        else:
            form = ShowRefuseFormUser(request.POST)
            context.update({'num_active_supplies': request.session['num_active_supplies']})

            if form.is_valid():
                postoffices = Postoffice.objects.filter(pk=user.postoffice_id.pk)
                params_for_opss.update({'postoffice__in': postoffices})

            else:
                messages.error(request, get_errors_form(form))



        date_s = request.POST.get('date_s', False)
        date_p = request.POST.get('date_p', False)

        if date_s:
            params_for_supplies.update({'date_sending__date__gte': date_s})

        if date_p:
            params_for_supplies.update({'date_sending__date__lte': date_p})


        parts_final = []
        if 'f_cartridges' in request.POST:
            opss = OPS.objects.filter(**params_for_opss)
            supplies_ops = Supply_OPS.objects.filter(**params_for_supplies, ops_recipient__in=opss, status_sending=True)
            parts_ops = Part_OPS.objects.filter(id_supply_ops__in=supplies_ops).order_by('cartridge')


            parts_ops_cartridges_unique = []
            for x in parts_ops:
                if x.cartridge not in parts_ops_cartridges_unique:
                    parts_ops_cartridges_unique.append(x.cartridge)



            for p in parts_ops_cartridges_unique:
                pfilter = parts_ops.filter(cartridge=p)
                total = 0
                mid = list()
                for pf in pfilter:
                    supid_user = pf.id_supply_ops.user_sender
                    mid_name = supid_user.middle_name if supid_user.middle_name else ""
                    user_name = "{0} {1} {2}".format(supid_user.last_name,
                                                     supid_user.first_name,
                                                     mid_name)


                    mid.append([pf.id_supply_ops.pk,
                                pf.id_supply_ops.ops_recipient.postoffice.postoffice_name,
                                pf.id_supply_ops.ops_recipient.index,
                                user_name,
                                formats.date_format(pf.id_supply_ops.date_sending, "DATE_FORMAT"),
                                pf.amount])
                    total += pf.amount

                parts_final.append([p.nomenclature, mid[:], total])

            request.session['sr_parts'] = parts_final

            return HttpResponseRedirect(reverse('show_refuse'))

        context.update({'title': 'Реестр списанных картриджей', 'form': form, 'user': user, 'sr_parts': parts_final})
        return render(request, 'showrefuse.html', context=context)


class AddOrder(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        user = request.user

        form_part_change = AutoorderFormPartChange()
        form_new_part = AutoorderFormNewPart()

        context = {'user': user,
                   'title': 'Заказ на поставку картриджей',
                   'num_active_supplies': request.session['num_active_supplies'],
                   'form': form_part_change,
                   'form_new_part': form_new_part}

        td = datetime.today()
        last_day = td.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
        month_year_for = last_day.strftime("%m%Y")
        order_active = AutoOrder.objects.filter(month_year_for=month_year_for, postoffice_autoorder=user.postoffice_id, status_sending=False).first()

        if order_active:
            autoorder_parts = Part_AutoOrder.objects.filter(id_autoorder=order_active)
            context.update({'autoorder_parts': autoorder_parts,
                            'autoorder_id': order_active.pk})

        return render(request, 'addorder.html', context=context)


    def post(self, request):

        user = request.user
        postoffice = user.postoffice_id
        context = {'title': 'Заказ на поставку картриджей', 'user': user}

        form_part_change = AutoorderFormPartChange(request.POST)
        form_new_part = AutoorderFormNewPart(request.POST)

        context.update({
            'num_active_supplies': request.session['num_active_supplies'],
            'form': form_part_change,
            'form_new_part': form_new_part})

        td = datetime.today()
        last_day = td.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
        first_day = td.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(days=last_day.day)
        month_year_for = last_day.strftime("%m%Y")


        if 'autoformer' in request.POST:

            order_active = AutoOrder.objects.filter(month_year_for=month_year_for, postoffice_autoorder=postoffice).first()

            if order_active == None:

                opss_postoffice = OPS.objects.filter(postoffice=user.postoffice_id)

                opss_supplies = Supply_OPS.objects.filter(ops_recipient__in=opss_postoffice,
                                                          status_sending=True,
                                                          date_sending__date__gte=first_day,
                                                          date_sending__date__lte=last_day)

                parts_supplies = Part_OPS.objects.filter(id_supply_ops__in=opss_supplies) \
                    .values(cid=F('cartridge__id'), printer=F('cartridge__printer_model')) \
                    .annotate(total_amount=Sum('amount')) \
                    .order_by('-total_amount')

                # create new AutoOrder
                new_autoorder = AutoOrder(postoffice_autoorder=postoffice,
                                          date_sending=datetime.now(),
                                          month_year_for=month_year_for,
                                          status_sending=False,
                                          user_autoorder=user)
                new_autoorder.save()

                # create new Parts auto order
                for part in parts_supplies:
                    Cartridge.objects.get(id=part['cid'])
                    new_part = Part_AutoOrder(cartridge=Cartridge.objects.get(id=part['cid']), amount=part['total_amount'], id_autoorder=new_autoorder)
                    new_part.save()

            else:
                if order_active.status_sending != True:
                    newparts = Part_AutoOrder.objects.filter(id_autoorder=order_active.pk, flag_n=True)
                    if newparts.exists():
                        newparts.delete()

                    parts_active_autoorder = Part_AutoOrder.objects.filter(id_autoorder=order_active.pk)

                    for p in parts_active_autoorder:
                        p.add_amount = 0
                        p.save()
                else:
                    messages.error(request, f'Автозаказ №{order_active.pk} на поставку картриджей был отправлен {order_active.date_sending.strftime("%d.%m.%Y в %H:%M:%S")}')

            return HttpResponseRedirect(reverse('add_order'))


        id_autopart = False
        id_autoorder = False
        for var in request.POST:
            if var.startswith('autopart'):
                id_autopart = var.split('_')[1]
            if var.startswith('newpart'):
                id_autoorder = var.split('_')[1]

        if id_autopart:
            if form_part_change.is_valid():
                part_active = Part_AutoOrder.objects.get(id=id_autopart)
                part_addamount = int(request.POST.get('amount'))
                part_addamount = 0 if part_addamount < 0 else part_addamount
                part_active.add_amount += part_addamount
                part_active.save()
            return HttpResponseRedirect(reverse('add_order'))


        if id_autoorder:
            if form_new_part.is_valid():
                nomenclature = request.POST.get('nomenclature_cartridge')
                amount_newpart = request.POST.get('amount_newpart')

                autoorder = AutoOrder.objects.get(pk=id_autoorder)
                cartridge = Cartridge.objects.get(nomenclature=nomenclature)

                part_ex = Part_AutoOrder.objects.filter(cartridge__nomenclature=nomenclature).first()
                if part_ex:
                    part_ex.add_amount += int(amount_newpart)
                    part_ex.save()
                else:
                    new_part = Part_AutoOrder(id_autoorder=autoorder,
                                              cartridge=cartridge,
                                              amount=amount_newpart,
                                              add_amount=0,
                                              flag_n=True)
                    new_part.save()

                return HttpResponseRedirect(reverse('add_order'))
            else:
                messages.error(request, get_errors_form(form_new_part))



        unsent_order = AutoOrder.objects.filter(month_year_for=month_year_for,
                                                postoffice_autoorder=postoffice,
                                                status_sending=False).first()

        if 'sendautoorder' in request.POST:
            num_parts = Part_AutoOrder.objects.filter(id_autoorder=unsent_order).count()
            if unsent_order and num_parts > 0:
                unsent_order.date_sending = datetime.now()
                unsent_order.status_sending = True
                unsent_order.save()
                messages.success(request, f'Автозаказ №{unsent_order.pk} на поставку картриджей принят')
            elif unsent_order and num_parts <= 0:
                messages.error(request, f'В автозаказе №{unsent_order.pk} нет ни одной позиции')
            else:
                messages.error(request, f'Автозаказ не сформирован')

            return HttpResponseRedirect(reverse('add_order'))


        # all autoorder parts active
        if unsent_order:
            autoorder_parts = Part_AutoOrder.objects.filter(id_autoorder=unsent_order)
            context.update({'autoorder_parts': autoorder_parts,
                            'autoorder_id': unsent_order.pk})


        return render(request, 'addorder.html', context=context)



class ShowOrders(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        user = request.user
        context = {'user': user}

        if user.role != '1':
            return render(request, 'main.html', context=context)


        all_orders = AutoOrder.objects.filter(status_sending=True)
        list_orders = []
        for order in all_orders:
            order_parts = Part_AutoOrder.objects.filter(id_autoorder=order)
            s = ''
            for part in order_parts:
                s += f'{part.cartridge.nomenclature}:{part.amount + part.add_amount}; '
            list_orders.append((order, s))
            list_orders = sorted(list_orders, key=lambda x: x[0].pk, reverse=True)

        context.update({'title': 'Заказы с почтамтов на поставку картриджей',
                   'orders': list_orders,
                   'num_active_orders': request.session['num_active_orders']})
        return render(request, 'showorders.html', context=context)

    def post(self, request):

        context = {}
        user = request.user
        context.update({'title': 'Заказы с почтамтов на поставку картриджей', 'num_active_orders': request.session['num_active_orders']})


        if 'tolist' in request.POST:
            return HttpResponseRedirect(reverse('show_orders'))

        id_autoorder = False
        id_supply_autoorder = False
        id_changepart = False
        for var in request.POST:
            if var.startswith('order'):
                id_autoorder = var.split('_')[1]
            if var.startswith('supplyby'):
                id_supply_autoorder = var.split('_')[1]
            if var.startswith('change'):
                id_changepart = var.split('_')[1]


        autoorder = None

        if id_autoorder:
            autoorder = AutoOrder.objects.get(pk=id_autoorder)
            autoorder.viewed = True
            autoorder.save()


        if id_supply_autoorder:
            # create supply
            autoorder = AutoOrder.objects.get(pk=id_supply_autoorder)
            postoffice_autoorder = autoorder.postoffice_autoorder
            autoparts_to_sup = Part_AutoOrder.objects.annotate(sum_amount=F('amount')+F('add_amount')).filter(id_autoorder=autoorder, sum_amount__gt=0)

            if not autoorder.workedout:
                # find sklad
                a_group = user.group
                a_sklad = Postoffice.objects.get(group=a_group, as_base=True)

                flag_validate = all([(part_autoorder.amount+part_autoorder.add_amount) <= State.objects.get(postoffice=a_sklad,
                                                                                                          cartridge=part_autoorder.cartridge).total_amount
                                     for part_autoorder in autoparts_to_sup])

                if flag_validate:
                    new_supply = Supply(postoffice_recipient=postoffice_autoorder.postoffice_name)
                    new_supply.save()

                    # create supply parts
                    for part_autoorder in autoparts_to_sup:
                        tot_amount = part_autoorder.amount+part_autoorder.add_amount

                        new_part = Part(id_supply=new_supply.pk,
                                        postoffice=new_supply.postoffice_recipient,
                                        cartridge=part_autoorder.cartridge.nomenclature,
                                        amount=tot_amount)


                        state = State.objects.get(postoffice=a_sklad, cartridge=part_autoorder.cartridge)
                        state.blocked_amount += tot_amount
                        state.total_amount -= tot_amount
                        state.save()

                        new_part.save()

                    autoorder.workedout = True
                    autoorder.save()

                    return HttpResponseRedirect(reverse('show_orders'))
                else:
                    messages.error(request,f'Количество картриджей в поставке не должно превышать количества картриджей на складе')
            else:
                messages.error(request, f'Поставка по данному заказу была создана раннее')


        if id_changepart:
            form_changeamount = ChangeAmountAutoorderForm(request.POST)
            changepart = Part_AutoOrder.objects.get(pk=id_changepart)
            autoorder = changepart.id_autoorder
            if form_changeamount.is_valid():
                p_amount = int(request.POST.get('new_amount'))

                # change amount active part
                changepart.add_amount = 0
                changepart.amount = p_amount
                changepart.save()

            else:
                messages.error(request, get_errors_form(form_changeamount))



        # global context
        autoparts = Part_AutoOrder.objects.filter(id_autoorder=autoorder)

        autoparts_list = []
        for autopart in autoparts:
            cartridge = autopart.cartridge

            active_group = user.postoffice_id.group
            sklad = Postoffice.objects.get(group=active_group, as_base=True)

            state_cart_sklad = State.objects.filter(postoffice=sklad, cartridge=cartridge)
            amount_sklad = state_cart_sklad.first().total_amount if state_cart_sklad.exists() else 0

            npart = (autopart, amount_sklad)

            autoparts_list.append(npart)

        request.session['num_active_orders'] = AutoOrder.objects.filter(status_sending=True,
                                                                        viewed=False).count()

        context.update({'autoorder': autoorder, 'autoparts_list': autoparts_list,
                        'num_active_orders': request.session['num_active_orders'],
                        'form_changeamount': ChangeAmountAutoorderForm()})
        # end global context



        return render(request, 'showorder.html', context=context)



class AddToStock(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        user = request.user

        form = AutoorderFormNewPart()

        context = {'title': 'Поставка на склад', 'user': user, 'form': form}


        list_supplies_stock = []
        supplies_stock = Supply_Stock.objects.filter(status_sending=False)
        for sup in supplies_stock:
            sup_parts = Part_Stock.objects.filter(supply_stock=sup)
            list_supplies_stock.append((sup, sup_parts))

        context.update({'list_supplies_stock': list_supplies_stock})


        return render(request, 'addtostock.html', context=context)



    def post(self, request):
        context = {}
        user = request.user
        context.update({'title': 'Поставка на склад',
                        'num_active_orders': request.session['num_active_orders']})

        form = AutoorderFormNewPart(request.POST)

        if 'addstock' in request.POST:
            new_supply_stock = Supply_Stock(user_sender=user)
            new_supply_stock.save()
            return HttpResponseRedirect(reverse('add_tostock'))

        id_supply_stock_newpart = False
        id_supply_stock_delsup = False
        id_supply_stock_delpart = False
        id_supply_stock_sendsup = False
        for var in request.POST:
            if var.startswith('newpartstock'):
                id_supply_stock_newpart = var.split('_')[1]
            if var.startswith('delsup'):
                id_supply_stock_delsup = var.split('_')[1]
            if var.startswith('delpart'):
                id_supply_stock_delpart = var.split('_')[1]
            if var.startswith('sendsup'):
                id_supply_stock_sendsup = var.split('_')[1]


        if id_supply_stock_newpart:
            if form.is_valid():
                nomenclature = request.POST.get('nomenclature_cartridge')
                amount = request.POST.get('amount_newpart')

                supply_stock = Supply_Stock.objects.get(pk=id_supply_stock_newpart)
                cartridge = Cartridge.objects.get(nomenclature=nomenclature)

                ex_part = Part_Stock.objects.filter(supply_stock=supply_stock, cartridge=cartridge).first()
                if ex_part:
                    ex_part.amount += int(amount)
                    ex_part.save()
                else:
                    new_part = Part_Stock(supply_stock=supply_stock, cartridge=cartridge, amount=amount)
                    new_part.save()

                return HttpResponseRedirect(reverse('add_tostock'))

            else:
                messages.error(request, get_errors_form(form), extra_tags=f'addpart_{id_supply_stock_newpart}')


        if id_supply_stock_delpart:
            delpart = Part_Stock.objects.get(pk=id_supply_stock_delpart)
            delpart.delete()


        if id_supply_stock_delsup:
            sup_todel = Supply_Stock.objects.get(pk=id_supply_stock_delsup)
            parts_todel = Part_Stock.objects.filter(supply_stock=sup_todel)
            for part in parts_todel:
                part.delete()
            sup_todel.delete()


        if id_supply_stock_sendsup:
            active_supply = Supply_Stock.objects.get(pk=id_supply_stock_sendsup)
            active_parts = Part_Stock.objects.filter(supply_stock=active_supply)

            if active_parts:
                sklad = Postoffice.objects.get(as_base=True)
                for part in active_parts:
                    active_state = State.objects.filter(postoffice=sklad, cartridge=part.cartridge).first()
                    if active_state:
                        active_state.total_amount += int(part.amount)
                        active_state.save()
                    else:
                        new_state = State(postoffice=sklad, cartridge=part.cartridge, total_amount=int(part.amount))
                        new_state.save()

                active_supply.user_sender = user
                active_supply.date_sending = datetime.now()
                active_supply.status_sending = True
                active_supply.save()
            else:
                messages.error(request, 'Невозможно отправить пустую поставку', extra_tags=f'sendsup_{id_supply_stock_sendsup}')

        list_supplies_stock = []
        supplies_stock = Supply_Stock.objects.filter(status_sending=False)
        for sup in supplies_stock:
            sup_parts = Part_Stock.objects.filter(supply_stock=sup)
            list_supplies_stock.append((sup, sup_parts))

        context.update({'list_supplies_stock': list_supplies_stock, 'form': form})



        return render(request, 'addtostock.html', context=context)


class MoveCartridges(View):

    def get(self, request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse('login'))

        user = request.user
        context = {'user': user}

        return render(request, 'movecartridges.html', context=context)



    def post(self, request):
        ...
